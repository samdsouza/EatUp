# Eat Up Driver File

import xlsxwriter
from openpyxl import load_workbook
import pandas

import xlrd





def fileLocation(fileName):
    baseDirectory = "/Users/samdsouza/Documents/EatUp/EatUp"
    fileLocation = baseDirectory + "/" + fileName
    return fileLocation

def openWorkbook(fileLocation):
    wb = xlrd.open_workbook(fileLocation)
    return wb


def returnCustomerID(workbook):
    sheet = workbook.sheet_by_index(0)
    sheet.cell_value(0,0)
    return sheet.cell_value(0,1)

def returnCustomerName(workbook):
    sheet = workbook.sheet_by_index(0)
    sheet.cell_value(0,0)
    return sheet.cell_value(1,1)

def extractInfo(workbook):
    sheet = workbook.sheet_by_index(0)
    sheet.cell_value(0,0)
    lengthSheet = sheet.nrows

    dates = []
    restaurants = []
    prices = []
    rewards = []

    for i in range(4,lengthSheet):
        #print(sheet.cell_value(i,0))
        dates.append(sheet.cell_value(i,0))
        restaurants.append(sheet.cell_value(i,1))
        prices.append(sheet.cell_value(i,2))
        rewards.append(sheet.cell_value(i,3))

    return dates, restaurants, prices, rewards


def searchRestaurant(workbook,searchRestaurant):
    sheet = workbook.sheet_by_index(0)
    sheet.cell_value(0,0)
    lengthSheet = sheet.nrows

    for i in range(4,lengthSheet):
        currentRestaurant = sheet.cell_value(i,1)
        if(searchRestaurant == currentRestaurant):
            return currentRestaurant
        else:
            print('Did Not Find Resteraunt')

def determineRewards(workbook,searchRestaurant):
    sheet = workbook.sheet_by_index(0)
    sheet.cell_value(0,0)
    lengthSheet = sheet.nrows

    for i in range(4,lengthSheet):
        currentRestaurant = sheet.cell_value(i,1)
        if(searchRestaurant == currentRestaurant):
            totalRewards =  sheet.cell_value(i,3)
            return totalRewards

def addNewVisit(fileDestination,restaurantName,date,totalSpent):
    # Before adding total spent, need to lookup resteraunt reward policy in resteraunt index
    # Determine the rewards earned and input into the individual resteraunt and customer excell sheet
    wb = load_workbook(filename = dest)
    worksheet = wb["Sheet1"]


    currentRow = worksheet.max_row
    dateCell = worksheet.cell(row = currentRow, column = 1)
    restaurantCell = worksheet.cell(row = currentRow, column = 2)
    spentCell = worksheet.cell(row = currentRow, column = 3)
    rewardsEarnedCell = worksheet.cell(row = currentRow, column = 4)
    rewardsClaimedCell = worksheet.cell(row = currentRow, column = 5)

    dateCell.value = date + "'"
    restaurantCell.value = restaurantName
    spentCell.value = totalSpent
    wb.save(dest)









# def writeNewRestaurant(filename, )





if __name__ == '__main__':
    print('Welcome to EatUp!')

    #restaurantIndex = xlsxwriter.Workbook('restaurantIndex.xlsx')
    #restaurantListings = restaurantIndex.add_worksheet()


    #restaurantListings.set_column('A:A', 20)
    #bold = restaurantIndex.add_format({'bold': True})

    #restaurantListings.write(1, 1, 'Resteraunt Name', bold)
    #restaurantListings.write(1,0,'Resteraunt ID', bold)

    #restaurantListings.write(2,0,1)
    #restaurantListings.write(2,1,'Chick Fil A')

    #restaurantListings.write(3,0,2)
    #restaurantListings.write(3,1, 'AeroTek Cafe')

    #restaurantIndex.close()


    fileLocation = fileLocation('customer1.xlsx')
    customer1 = openWorkbook(fileLocation)
    customerID = returnCustomerID(customer1)
    customerName = returnCustomerName(customer1)
    results = extractInfo(customer1)
    customer2Dates = results[0]
    customer2Restaurants = results[1]
    customer2Prices = results[2]
    customer2Rewards = results[3]



    print('Before Search Call')
    currentRewards = determineRewards(customer1, 'Via Perla')

    #restaurantIndex = xlsxwriter.Workbook('restaurantIndex.xlsx')
    dest = "/Users/samdsouza/Documents/EatUp/EatUp/customer1.xlsx"

    #addNewVisit(dest,"Pizzeria Locale","4/2/2020","55.76")
    #wb = load_workbook(filename = dest)
    #worksheet = wb["Sheet1"]

    #c = worksheet.cell(row = 13, column = 1)
    #c.value = 'Testing 25'
    #row_count = worksheet.max_row
    #print(row_count)

    #wb.save(dest)










    #ref_workbook= openpyxl.load_workbook('resterauntIndex.xlsx')

    #customer2 = pandas.read_excel('customer2.xlsx')
    #print(customer2)

    #loc = ("/Users/samdsouza/Documents/EatUp/EatUp/customer2.xlsx")

    #wb = xlrd.open_workbook(loc)
    #sheet = wb.sheet_by_index(0)

    #sheet.cell_value(0,0)

    #for i in range(sheet.nrows):
    #    print(sheet.cell_value(i,0))
