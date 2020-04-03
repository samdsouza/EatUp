import xlsxwriter
from openpyxl import load_workbook
import pandas

import xlrd




def main():
    #Framework for demo functionality including all menu options
    #display upon entry and entry only:
    print("_________________ Welcome to Eatup! _________________")
    print("To exit type exit,")
    #Recieve user input and trasfer to correct menu, at this point this menu will be the only one displayed until exit
    command = "init"
    while command != "exit":
        print("_______________________ Menu ________________________\n")

        print("1. Buisness")
        print("2. Customer\n")

        command = input()
        command.strip()
        print("Command = ", command)
        if command == "Buisness":
            while command != "back":
                buisnessMenu()
                command = input()
                if command == "Setup":
                    setupBuisness()
                if command == "Manage":
                    print("Enter Buisness ID:")
                    command = input()
                    login_success = 0
                    login_success, buisnessName = loginBuisness(command)
                    while login_success != 1:
                        print("Buisness ID not found in database, Please try again\n")
                        command = input()
                        login_success, buisnessName = loginBuisness(command)
                    manageMenu(buisnessName)

        elif command == "Customer":
            while command != "back":
                customerMenu()
                command = input()
                if command == "New user":
                    setupUser()

                if command == "Login":
                    print("Enter Customer ID:")
                    customerID = input()
                    login_success = 0
                    login_success, customerFirstName, customerLastName = loginCustomer(customerID)
                    while login_success != 1:
                        print("Customer ID not found in database, Please try again")
                        customerID = input()
                        login_success, customerFirstName, customerLastName = loginCustomer(customerID)
                    print('Welcome Back' , customerFirstName.strip(), customerLastName.strip(), "!")
                    customerRewardsMenu(customerID, customerFirstName.strip(),customerLastName.strip())

        elif command == "exit":
            return

        else:
            print("Invaild Command")

    return











def fileLocation(fileName):
    baseDirectory = "/Users/samdsouza/Documents/EatUp/EatUp"
    fileLocation = baseDirectory + "/" + fileName
    return fileLocation

def openWorkbook(fileLocation):
    wb = xlrd.open_workbook(fileLocation)
    return wb


def returnCustomerID(workbook):
    sheet = workbook.sheet_by_index(0)
    sheet.cell_value(0,0)
    return sheet.cell_value(0,1)

def returnCustomerName(workbook):
    sheet = workbook.sheet_by_index(0)
    sheet.cell_value(0,0)
    return sheet.cell_value(1,1)

def extractInfo(workbook):
    sheet = workbook.sheet_by_index(0)
    sheet.cell_value(0,0)
    lengthSheet = sheet.nrows

    dates = []
    restaurants = []
    prices = []
    rewards = []

    for i in range(4,lengthSheet):
        #print(sheet.cell_value(i,0))
        dates.append(sheet.cell_value(i,0))
        restaurants.append(sheet.cell_value(i,1))
        prices.append(sheet.cell_value(i,2))
        rewards.append(sheet.cell_value(i,3))

    return dates, restaurants, prices, rewards


def searchRestaurant(workbook,searchRestaurant):
    sheet = workbook.sheet_by_index(0)
    sheet.cell_value(0,0)
    lengthSheet = sheet.nrows

    for i in range(4,lengthSheet):
        currentRestaurant = sheet.cell_value(i,1)
        if(searchRestaurant == currentRestaurant):
            return currentRestaurant
        else:
            print('Did Not Find Resteraunt')

def determineRewards(workbook,searchRestaurant):
    #print('made it into rewards function')
    #print('search resteraunt = ', searchRestaurant)
    sheet = workbook.sheet_by_index(0)
    sheet.cell_value(0,0)
    lengthSheet = sheet.nrows

    for i in range(4,lengthSheet):
        currentRestaurant = sheet.cell_value(i,1)
        #print('Current resteraunt = ', currentRestaurant)
        if(searchRestaurant.strip() == currentRestaurant.strip()):
            totalRewards =  sheet.cell_value(i,3)
            return totalRewards


def listResteraunts(workbook, printFlag):
    sheet = workbook.sheet_by_index(0)
    sheet.cell_value(0,0)
    lengthSheet = sheet.nrows
    resterauntList = []

    for i in range(4,lengthSheet):
        currentRestaurant = sheet.cell_value(i,1)
        resterauntList.append(currentRestaurant.strip())
        if printFlag == True:
            print(currentRestaurant)
    return resterauntList


def addNewVisit(fileDestination,restaurantName,date,totalSpent):
    # Before adding total spent, need to lookup resteraunt reward policy in resteraunt index
    # Determine the rewards earned and input into the individual resteraunt and customer excell sheet
    wb = load_workbook(filename = fileDestination)
    worksheet = wb["Sheet1"]


    currentRow = worksheet.max_row
    dateCell = worksheet.cell(row = currentRow, column = 1)
    restaurantCell = worksheet.cell(row = currentRow, column = 2)
    spentCell = worksheet.cell(row = currentRow, column = 3)
    rewardsEarnedCell = worksheet.cell(row = currentRow, column = 4)
    rewardsClaimedCell = worksheet.cell(row = currentRow, column = 5)

    rewardsEarnedCell.value = 1
    dateCell.value = date + "'"
    restaurantCell.value = restaurantName
    spentCell.value = totalSpent
    wb.save(fileDestination)


def determineResterauntIndex(workbook, searchRestaurant):
    print("In resteraunt Index")
    sheet = workbook.sheet_by_index(0)
    sheet.cell_value(0,0)
    lengthSheet = sheet.nrows
    #print("Sheet Length = ", lengthSheet)

    for i in range(4,lengthSheet):
        currentRestaurant = sheet.cell_value(i,1)
        #print("currentResteraunt = ",currentRestaurant )
        if(str(searchRestaurant.strip()) == str(currentRestaurant.strip())):
            totalCost = sheet.cell_value(i,2)
            #print("Total Cost = ", totalCost)
            return i, totalCost


def updateVisit(fileDestination,restaurantName,date,totalSpent,resterauntIndex, currentCost):

    wb = load_workbook(filename = fileDestination)
    worksheet = wb["Sheet1"]

    spentCell = worksheet.cell(row = resterauntIndex + 1, column = 3)
    spentCell.value = currentCost + float(totalSpent)
    wb.save(fileDestination)





#Print menu definitions
def buisnessMenu():
    print("___________________  Buisness Menu ____________________\n")
    print("1. Setup")
    print("2. Manage\n")
    return

#manageMenu handling
def manageMenu(buisnessName):
    inp = "init"
    while inp != "back":
        print("____________________ ", buisnessName, " _____________________\n")
        print("1. Add rewards")
        print("2. Remove rewards")
        print("3. View rewards")
        inp = input()

        if inp == "Add rewards":
            addRewards()
        if inp == "Remove rewards":
            removeRewards()
        if inp == "View rewards":
            displayRewards()
    return

#customer menu function
def customerMenu():
    print("___________________  Customer Menu _______________________\n")
    print("1. New user")
    print("2. Login\n")
    return

#Main customer functionality
def customerRewardsMenu(customerID, customerFirstName, customerLastName):
    #print('Made it in customer Rewards ')
    inp = "start"
    while inp != "back":
        print("\n")
        print("____________________ ", customerFirstName, customerLastName.strip(),"'s", "Rewards_________________")
        print("1. Rewards Progress")
        print("2. Input Rewards")

        inp = input()
        inp.strip()
        if inp == "Rewards Progress":
            rewardsProgress(customerID)
        if inp == "Input Rewards":
            inputRewards(customerID)
        if inp == "Redeem Rewards":
            redeemRewards()
    return

#placeholder
def loginCustomer(command):
    signal = 1
    filename = 'customerIndex.xlsx'
    fileLocationOutput = fileLocation(filename)
    customerIndexBook = openWorkbook(fileLocationOutput)
    sheet = customerIndexBook.sheet_by_index(0)
    sheet.cell_value(0,0)
    lengthSheet = sheet.nrows
    #print('Length of Sheet =' , lengthSheet)

    for i in range(2,lengthSheet):
        #('Starting Loop')
        currentID = sheet.cell_value(i,0)
        #print('Current ID = ', currentID)
        if(command == str(int(currentID))):
            customerFirstName =  sheet.cell_value(i,1)
            customerLastName = sheet.cell_value(i,2)
            return signal, customerFirstName, customerLastName

    signal = 0
    customerFirstName = 'NA'
    customerLastName = 'NA'
    return signal, customerFirstName, customerLastName



#placeholder
def rewardsProgress(customerID):
    filename = 'customer' + customerID + '.xlsx'
    fileLocationOutput = fileLocation(filename)
    customerWorkbook = openWorkbook(fileLocationOutput)

    print("\n")
    print('Favorite Resteraunts')
    resterauntList = listResteraunts(customerWorkbook, True)


    print("\n")
    print('Enter Resteraunt Name')
    resterauntName = input()

    resterauntRewards = determineRewards(customerWorkbook,resterauntName)

    print('Total Rewards for ', resterauntName, 'is ', str(int(resterauntRewards)))

    return

#placeholder
def inputRewards(customerID):

    filename = 'customer' + customerID + '.xlsx'
    fileLocationOutput = fileLocation(filename)
    customerWorkbook = openWorkbook(fileLocationOutput)

    print("\n")

    print("Enter Resteraunt Name")
    resterauntName = input()
    resterauntName.strip()

    print("Ente Date of Purchase")
    date = input()
    date.strip()

    print("Enter Total Cost Spent")
    totalCost = input()
    totalCost.strip()

    resterauntListing = listResteraunts(customerWorkbook, False)

    #print(resterauntListing)



    if resterauntName in resterauntListing:
        print('In current Resteraunt')
        resterauntIndex, currentCost = determineResterauntIndex(customerWorkbook,resterauntName)
        updateVisit(filename,resterauntName,date,totalCost,resterauntIndex, currentCost)
        #print(this)
    else:
        print("New Resteraunt")
        addNewVisit(filename,resterauntName,date,totalCost)


    return

#placeholder
def redeemRewards():
    return

#placeholder
def addRewards():
    return

#placeholder
def removeRewards():
    return

#placeholder
def displayRewards():
    return

#placeholder
def editRewards():
    return

def setupUser():
    return

#placeholder for buisness Setup
def setupBuisness():
    return

#placeholder for buisness login
def loginBuisness(command):
    signal = 1
    filename = 'restaurantIndex.xlsx'
    fileLocationOutput = fileLocation(filename)
    buisnessIndexBook = openWorkbook(fileLocationOutput)
    sheet = buisnessIndexBook.sheet_by_index(0)
    sheet.cell_value(0,0)
    lengthSheet = sheet.nrows
    #print('Length of Sheet =' , lengthSheet)

    for i in range(2,lengthSheet):
        #('Starting Loop')
        currentID = sheet.cell_value(i,0)
        #print('Current ID = ', currentID)
        if(command == str(int(currentID))):
            buisnessName =  sheet.cell_value(i,1)
            return signal, buisnessName

    signal = 0
    buisnessName = 'NA'
    return signal, buisnessName

if __name__ == "__main__":
    main()
